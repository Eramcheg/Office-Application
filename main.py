import tkinter as tk
import openpyxl
from PIL import Image
import os
from tkinter import filedialog
from openpyxl.styles import Alignment
from tkinter import ttk


class MainClass:
    def __init__(self):
        self.filepath = ""
        self.dirImages = []
        self.columnArticles = 'B'
        self.columnImages = 'A'
        self.selectedSheet = ''

    def startApp(self):
        root = tk.Tk()

        # all possible columns
        alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
                    'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK',
                    'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ', 'BA',
                    'BB', 'BC', 'BD', 'BE', 'BF', 'BG', 'BH', 'BI', 'BJ', 'BK', 'BL', 'BM', 'BN', 'BO', 'BP', 'BQ',
                    'BR', 'BS', 'BT', 'BU', 'BV', 'BW', 'BX', 'BY', 'BZ', 'CA', 'CB', 'CC', 'CD', 'CE', 'CF', 'CG',
                    'CH', 'CI', 'CJ', 'CK', 'CL', 'CM', 'CN', 'CO', 'CP', 'CQ', 'CR', 'CS', 'CT', 'CU', 'CV', 'CW',
                    'CX', 'CY', 'CZ', 'DA', 'DB', 'DC', 'DD', 'DE', 'DF', 'DG', 'DH', 'DI', 'DJ', 'DK', 'DL', 'DM',
                    'DN', 'DO', 'DP', 'DQ', 'DR', 'DS', 'DT', 'DU', 'DV', 'DW', 'DX', 'DY', 'DZ', 'EA', 'EB', 'EC',
                    'ED', 'EE', 'EF', 'EG', 'EH', 'EI', 'EJ', 'EK', 'EL', 'EM', 'EN', 'EO', 'EP', 'EQ', 'ER', 'ES',
                    'ET', 'EU', 'EV', 'EW', 'EX', 'EY', 'EZ', 'FA', 'FB', 'FC', 'FD', 'FE', 'FF', 'FG', 'FH', 'FI',
                    'FJ', 'FK', 'FL', 'FM', 'FN', 'FO', 'FP', 'FQ', 'FR', 'FS', 'FT', 'FU', 'FV', 'FW', 'FX', 'FY',
                    'FZ', 'GA', 'GB', 'GC', 'GD', 'GE', 'GF', 'GG', 'GH', 'GI', 'GJ', 'GK', 'GL', 'GM', 'GN', 'GO',
                    'GP', 'GQ', 'GR', 'GS', 'GT', 'GU', 'GV', 'GW', 'GX', 'GY', 'GZ', 'HA', 'HB', 'HC', 'HD', 'HE',
                    'HF', 'HG', 'HH', 'HI', 'HJ', 'HK', 'HL', 'HM', 'HN', 'HO', 'HP', 'HQ', 'HR', 'HS', 'HT', 'HU',
                    'HV', 'HW', 'HX', 'HY', 'HZ', 'IA', 'IB', 'IC', 'ID', 'IE', 'IF', 'IG', 'IH', 'II', 'IJ', 'IK',
                    'IL', 'IM', 'IN', 'IO', 'IP', 'IQ', 'IR', 'IS', 'IT', 'IU', 'IV', 'IW', 'IX', 'IY', 'IZ', 'JA',
                    'JB', 'JC', 'JD', 'JE', 'JF', 'JG', 'JH', 'JI', 'JJ', 'JK', 'JL', 'JM', 'JN', 'JO', 'JP', 'JQ',
                    'JR', 'JS', 'JT', 'JU', 'JV', 'JW', 'JX', 'JY', 'JZ', 'KA', 'KB', 'KC', 'KD', 'KE', 'KF', 'KG',
                    'KH', 'KI', 'KJ', 'KK', 'KL', 'KM', 'KN', 'KO', 'KP', 'KQ', 'KR', 'KS', 'KT', 'KU', 'KV', 'KW',
                    'KX', 'KY', 'KZ', 'LA', 'LB', 'LC', 'LD', 'LE', 'LF', 'LG', 'LH', 'LI', 'LJ', 'LK', 'LL', 'LM',
                    'LN', 'LO', 'LP', 'LQ', 'LR', 'LS', 'LT', 'LU', 'LV', 'LW', 'LX', 'LY', 'LZ', 'MA', 'MB', 'MC',
                    'MD', 'ME', 'MF', 'MG', 'MH', 'MI', 'MJ', 'MK', 'ML', 'MM', 'MN', 'MO', 'MP', 'MQ', 'MR', 'MS',
                    'MT', 'MU', 'MV', 'MW', 'MX', 'MY', 'MZ', 'NA', 'NB', 'NC', 'ND', 'NE', 'NF', 'NG', 'NH', 'NI',
                    'NJ', 'NK', 'NL', 'NM', 'NN', 'NO', 'NP', 'NQ', 'NR', 'NS', 'NT', 'NU', 'NV', 'NW', 'NX', 'NY',
                    'NZ', 'OA', 'OB', 'OC', 'OD', 'OE', 'OF', 'OG', 'OH', 'OI', 'OJ', 'OK', 'OL', 'OM', 'ON', 'OO',
                    'OP', 'OQ', 'OR', 'OS', 'OT', 'OU', 'OV', 'OW', 'OX', 'OY', 'OZ', 'PA', 'PB', 'PC', 'PD', 'PE',
                    'PF', 'PG', 'PH', 'PI', 'PJ', 'PK', 'PL', 'PM', 'PN', 'PO', 'PP', 'PQ', 'PR', 'PS', 'PT', 'PU',
                    'PV', 'PW', 'PX', 'PY', 'PZ', 'QA', 'QB', 'QC', 'QD', 'QE', 'QF', 'QG', 'QH', 'QI', 'QJ', 'QK',
                    'QL', 'QM', 'QN', 'QO', 'QP', 'QQ', 'QR', 'QS', 'QT', 'QU', 'QV', 'QW', 'QX', 'QY', 'QZ', 'RA',
                    'RB', 'RC', 'RD', 'RE', 'RF', 'RG', 'RH', 'RI', 'RJ', 'RK', 'RL', 'RM', 'RN', 'RO', 'RP', 'RQ',
                    'RR', 'RS', 'RT', 'RU', 'RV', 'RW', 'RX', 'RY', 'RZ', 'SA', 'SB', 'SC', 'SD', 'SE', 'SF', 'SG',
                    'SH', 'SI', 'SJ', 'SK', 'SL', 'SM', 'SN', 'SO', 'SP', 'SQ', 'SR', 'SS', 'ST', 'SU', 'SV', 'SW',
                    'SX', 'SY', 'SZ', 'TA', 'TB', 'TC', 'TD', 'TE', 'TF', 'TG', 'TH', 'TI', 'TJ', 'TK', 'TL', 'TM',
                    'TN', 'TO', 'TP', 'TQ', 'TR', 'TS', 'TT', 'TU', 'TV', 'TW', 'TX', 'TY', 'TZ', 'UA', 'UB', 'UC',
                    'UD', 'UE', 'UF', 'UG', 'UH', 'UI', 'UJ', 'UK', 'UL', 'UM', 'UN', 'UO', 'UP', 'UQ', 'UR', 'US',
                    'UT', 'UU', 'UV', 'UW', 'UX', 'UY', 'UZ', 'VA', 'VB', 'VC', 'VD', 'VE', 'VF', 'VG', 'VH', 'VI',
                    'VJ', 'VK', 'VL', 'VM', 'VN', 'VO', 'VP', 'VQ', 'VR', 'VS', 'VT', 'VU', 'VV', 'VW', 'VX', 'VY',
                    'VZ', 'WA', 'WB', 'WC', 'WD', 'WE', 'WF', 'WG', 'WH', 'WI', 'WJ', 'WK', 'WL', 'WM', 'WN', 'WO',
                    'WP', 'WQ', 'WR', 'WS', 'WT', 'WU', 'WV', 'WW', 'WX', 'WY', 'WZ', 'XA', 'XB', 'XC', 'XD', 'XE',
                    'XF', 'XG', 'XH', 'XI', 'XJ', 'XK', 'XL', 'XM', 'XN', 'XO', 'XP', 'XQ', 'XR', 'XS', 'XT', 'XU',
                    'XV', 'XW', 'XX', 'XY', 'XZ', 'YA', 'YB', 'YC', 'YD', 'YE', 'YF', 'YG', 'YH', 'YI', 'YJ', 'YK',
                    'YL', 'YM', 'YN', 'YO', 'YP', 'YQ', 'YR', 'YS', 'YT', 'YU', 'YV', 'YW', 'YX', 'YY', 'YZ', 'ZA',
                    'ZB', 'ZC', 'ZD', 'ZE', 'ZF', 'ZG', 'ZH', 'ZI', 'ZJ', 'ZK', 'ZL', 'ZM', 'ZN', 'ZO', 'ZP', 'ZQ',
                    'ZR', 'ZS', 'ZT', 'ZU', 'ZV', 'ZW', 'ZX', 'ZY', 'ZZ']
        root.geometry("600x600")
        root.title("Main")
        root.configure(background="black")

        # grid layout
        buttonFrame = tk.Frame(root, background='light gray')
        buttonFrame.columnconfigure(0, weight=1)
        buttonFrame.columnconfigure(1, weight=1)

        buttonFrame.pack(pady=160, padx=0)

        # label, combobox list
        labelSheets=tk.Label(buttonFrame,text="Select a sheet of excel file for next operations", font=('Arial', 15), background='light gray')
        labelSheets.grid(row=0, column=1, columnspan=2, sticky=tk.E+tk.W)

        # label, has been excel file selected
        labelExcel = tk.Label(buttonFrame, text="          File is not selected         ", foreground='red',
                              font=('Arial', 10))
        labelExcel.grid(row=1, column=0)

        # label, has been folder selected?
        labelImage = tk.Label(buttonFrame, text="        Folder is not selected       ", foreground='red',
                              font=('Arial', 10))
        labelImage.grid(row=3, column=0)

        # label, is button insert pressed?
        labelInsert = tk.Label(buttonFrame, text="Click Insert to insert images into excel file", foreground='red',
                               font=('Arial', 11))
        labelInsert.grid(row=5, column=0, sticky=tk.E + tk.W, columnspan=3)

        # combobox excel sheets
        sheets = ttk.Combobox(buttonFrame, values=[])
        sheets.grid(row=1, column=1, columnspan=2)
        sheets.bind('<<ComboboxSelected>>', self.sheetChanged)

        # button select excel file
        buttonExcel = tk.Button(buttonFrame, text="Load excel file", font=('Arial', 17),
                                command=lambda: self.openFile(labelExcel, labelInsert, sheets))
        buttonExcel.grid(row=0, column=0, sticky=tk.E + tk.W)

        # button select image folder
        buttonImage = tk.Button(buttonFrame, text="Load image folder", font=('Arial', 17), width=10,
                                command=lambda: self.openImage(labelImage, labelInsert))
        buttonImage.grid(row=2, column=0, sticky=tk.E + tk.W)

        # button insert
        buttonSave = tk.Button(buttonFrame, text="Insert", font=('Arial', 19),
                               command=lambda: self.saveImage(labelInsert))
        buttonSave.grid(row=4, column=0, sticky=tk.E + tk.W, columnspan=3)

        # instructions label
        labelArticles = tk.Label(buttonFrame, text="Select the excel column\n"
                                                   "where the codes are located", font=('Arial', 11),
                                 background='light gray')
        labelArticles.grid(row=2, column=1)

        # instructions label
        labelImages = tk.Label(buttonFrame, text="Select the excel column\n"
                                                 "where the images will be located", font=('Arial', 11),
                               background='light gray')
        labelImages.grid(row=2, column=2)





        # combobox articles column
        columnArticles = ttk.Combobox(buttonFrame, values=alphabet, font=('Arial', 12), background='light gray')
        columnArticles.grid(row=3, column=1)
        columnArticles.set("Column with Numbers")

        columnArticles.bind('<<ComboboxSelected>>', self.articlesChanged)

        # combobox image column
        columnImages = ttk.Combobox(buttonFrame, values=alphabet, font=('Arial', 12), background='black')
        columnImages.set("Column for images")
        columnImages.grid(row=3, column=2)

        columnImages.bind('<<ComboboxSelected>>', self.imagesChanged)

        # quit button
        quit = tk.Button(buttonFrame, text="Quit", font=('Arial', 19), command=self.quitcode)
        quit.grid(row=6, column=0, sticky=tk.E + tk.W, columnspan=3)

        root.mainloop()

    def sheetChanged(self, event):
        self.selectedSheet= event.widget.get()
        print(event.widget.get())


    def articlesChanged(self, event):
        self.columnArticles = event.widget.get()
        print(self.columnArticles)

    def imagesChanged(self, event):
        self.columnImages = event.widget.get()
        print(self.columnImages)

    def openFile(self, label, labelInsert,combobox):
        tempdir = filedialog.askopenfilename(initialdir="/", title="Select An Excel File", filetypes=(
            ("excel files", "*.xlsx"), ("All files", "*.*")))
        if len(tempdir) > 0:
            self.filepath = tempdir
            label.configure(text="      File has been selected      ", foreground='green')
            labelInsert.configure(text="Click Insert to insert images into excel file", foreground='red',
                                  font=('Arial', 11))

            arr_of_sheets= (openpyxl.load_workbook(tempdir)).sheetnames
            combobox.configure(values=arr_of_sheets)
            combobox.set(arr_of_sheets[0])
            combobox.update()
            combobox.configure()

        label.update()
        labelInsert.update()
        print(tempdir)

    def openImage(self, label, labelInsert):
        tempdir = filedialog.askdirectory(initialdir="/", title="Select An Image Folder")
        if len(tempdir) > 0:
            self.imagePath = tempdir
            label.configure(text="    Folder has been selected    ", foreground='green')
            labelInsert.configure(text="Click Insert to insert images into excel file", foreground='red',
                                  font=('Arial', 11))
            label.update()
            labelInsert.update()
            self.dirImages = os.listdir(tempdir)
            self.imagePath = tempdir
            print(tempdir)

    def saveImage(self, label):

        if self.filepath != '' and self.dirImages != [] and self.selectedSheet!='':
            try:
                start = 1
                finish = len(self.dirImages)
                for i in range(start, finish + 1):

                    # opening workbook
                    workbook = openpyxl.load_workbook(self.filepath)

                    # opening sheet 1
                    worksheet = workbook["List1"]

                    # unique number of product in cell
                    artical_number = str(worksheet[self.columnArticles + str(i)].value)


                    for j in range(finish):
                        if artical_number in self.dirImages[j]:
                            # opening an image
                            img = Image.open(str(self.imagePath + "/" + self.dirImages[j]))
                            img1 = openpyxl.drawing.image.Image(img)

                            # Set new image size
                            img1.width = 200 * img1.width / img1.height
                            img1.height = 200

                            # Set row and column
                            worksheet.row_dimensions[i].height = int(150)

                            # width and height
                            worksheet.column_dimensions[self.columnImages].width = 44.6

                            cell = worksheet[self.columnImages + str(i)]
                            cell.alignment = Alignment(horizontal='right')
                            worksheet.add_image(img1, self.columnImages + str(i))  # Adding image to worksheet

                            workbook.save(self.filepath)  # Saving a document

                            break

                label.configure(text="File data has been updated!", foreground='green')
                label.update()
                print("Program is successful")

            except:
                label.configure(text="Something went wrong", foreground='orange')
                label.update()

    def quitcode(self):
        quit()


m = MainClass()
m.startApp()
